import csv
import logging
import os
import re
import shutil
import time
import win32com.client as win32

import yaml
from openpyxl.reader.excel import load_workbook

from typing import List
import _testconfig as tc
from datetime import datetime, timedelta

import pandas as pd
from dateutil.relativedelta import relativedelta
from selenium.webdriver.common.by import By


def contains_chinese(text):
    return any('\u4e00' <= char <= '\u9fff' for char in text)


class WebDriverUtils(object):
    @staticmethod
    def translate(value: str) -> str:
        return "translate({}, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz')".format(value)

    @staticmethod
    def transform_xpath_locator(*loc):
        loc_x = (By.XPATH, loc[0]) if len(loc) == 1 and isinstance(loc[0], str) and (loc[0][0] in ('/', '.')) else loc
        loc_x = (By.XPATH, "." + loc_x[1]) if loc_x[0] == By.XPATH and loc_x[1][0] != "." else loc_x
        return loc_x


class ReUtils(object):
    ptn_currency = r"[¥,$]([0-9]{1,3}(,[0-9]{3})*)(.[0-9]{1,2})?"

    @staticmethod
    def is_email(email):
        ptn_email = r'([A-Za-z0-9]+[.-_])*[A-Za-z0-9]+@[A-Za-z0-9-]+(\.[A-Z|a-z]{2,})+'
        pattern = re.compile(ptn_email)
        return True if re.fullmatch(pattern, email) else False


class DateTimeUtils(object):
    str_dt_format = '%Y/%m/%dT%H:%M:%S+00:00'
    str_date_format_slash = '%Y/%m/%d'
    str_date_format_dash = '%Y-%m-%d'
    date_pattern = r"^\d{4}[-/]\d{2}[-/]\d{2}$"

    @staticmethod
    def current_datetime_string(short_year=True, include_ms=False):
        now = datetime.now()
        year = str(now.year) if not short_year else str(now.year)[-2:]
        month = str(now.month) if now.month > 9 else f'0{now.month}'
        day = str(now.day) if now.day > 9 else f'0{now.day}'
        hour = str(now.hour) if now.hour > 9 else f'0{now.hour}'
        minute = str(now.minute) if now.minute > 9 else f'0{now.minute}'
        second = str(now.second) if now.second > 9 else f'0{now.second}'
        microsecond = str(now.microsecond) if include_ms else ''
        return f'{year}{month}{day}{hour}{minute}{second}{microsecond}'

    @staticmethod
    def add_year(delta, origin_date=None):
        origin_date = origin_date if not origin_date else datetime.now()
        if isinstance(origin_date, str):
            origin_date = datetime.strptime(origin_date, DateTimeUtils.str_dt_format)
        return origin_date + relativedelta(years=delta)

    @staticmethod
    def add_month(delta, origin_date=None):
        origin_date = origin_date if not origin_date else datetime.now()
        if isinstance(origin_date, str):
            origin_date = datetime.strptime(origin_date, DateTimeUtils.str_dt_format)
        return origin_date + relativedelta(months=delta)

    @staticmethod
    def add_week(delta, origin_date=None):
        origin_date = origin_date if not origin_date else datetime.now()
        if isinstance(origin_date, str):
            origin_date = datetime.strptime(origin_date, DateTimeUtils.str_dt_format)
        return origin_date + relativedelta(weeks=delta)

    @staticmethod
    def add_day(delta, origin_date=None):
        origin_date = origin_date if not origin_date else datetime.now()
        if isinstance(origin_date, str):
            origin_date = datetime.strptime(origin_date, DateTimeUtils.str_dt_format)
        return origin_date + relativedelta(days=delta)

    @staticmethod
    def add_hour(delta, origin_date=None):
        origin_date = origin_date if not origin_date else datetime.now()
        if isinstance(origin_date, str):
            origin_date = datetime.strptime(origin_date, DateTimeUtils.str_dt_format)
        return origin_date + relativedelta(hours=delta)

    @staticmethod
    def add_minute(delta, origin_date=None):
        origin_date = origin_date if not origin_date else datetime.now()
        if isinstance(origin_date, str):
            origin_date = datetime.strptime(origin_date, DateTimeUtils.str_dt_format)
        return origin_date + relativedelta(minutes=delta)

    @staticmethod
    def min_date(*days):
        if len(days) < 2:
            logging.warning('at least 2 datetime objects are required to compare')
            return None
        else:
            days_list = [datetime.fromisoformat(d) if isinstance(d, str) else d for d in days]
            days_list.sort()
            return days_list[0]

    @staticmethod
    def max_date(*days):
        if len(days) < 2:
            logging.warning('at least 2 datetime objects are required to compare')
            return None
        else:
            days_list = [datetime.fromisoformat(d) if isinstance(d, str) else d for d in days]
            days_list.sort()
            return days_list[-1]

    @staticmethod
    def validate_date(date_string, date_format=str_date_format_slash):
        try:
            datetime.strptime(date_string, date_format)
            return True
        except ValueError:
            return False

    @staticmethod
    def validate_date_format(date_string):
        if re.match(DateTimeUtils.date_pattern, date_string):
            return True
        else:
            return False


class CsvUtil(object):
    @classmethod
    def read(cls, filepath, start_row_num_including_header=0, end_row=-1, *columns):
        header, content, header_indexes = [], [], []
        with open(filepath) as csv_file:
            reader = csv.reader(csv_file)
            end_row = reader.line_num - 1 if end_row == -1 or end_row >= reader.line_num else end_row
            cursor = 0
            for row in reader:
                if cursor == 0:
                    header = [col_name for col_name in columns and col_name in row]
                    header_indexes = [row.index(col_name) for col_name in row]
                if start_row_num_including_header < cursor <= end_row:
                    content.append([row[i] for i in header_indexes])
                cursor += 1
        return header, content

    @classmethod
    def read_header(cls, filepath):
        with open(filepath) as csv_file:
            reader = csv.reader(csv_file)
            for row in reader:
                return row

    @classmethod
    def read_to_dict(cls, filepath, start_row_num_including_header=0, end_row=-1, *columns):
        header = list(*columns)
        i_header, content = [], []
        with open(filepath) as csv_file:
            reader = csv.reader(csv_file)
            # end_row = reader.line_num - 1 if end_row == -1 or end_row >= reader.line_num else end_row
            cursor = 0
            for row in reader:
                if cursor == 0:
                    header = [col_name for col_name in header if col_name in row]
                    i_header = [row.index(col_name) for col_name in row]
                if cursor > start_row_num_including_header:
                    content.append({header[i_header.index(i)]: row[i] for i in i_header})
                if end_row != -1 and cursor >= end_row:
                    break
                cursor += 1
        return header, content

    @classmethod
    def read_all_to_list(cls, filepath):
        header, content = cls.read(filepath)
        content.insert(0, header)
        return content

    @classmethod
    def read_all_to_list_without_header(cls, filepath):
        header, content = cls.read(filepath, 1)
        return content

    @classmethod
    def read_all_to_dict(cls, filepath):
        header = cls.read_header(filepath)
        header, content = cls.read_to_dict(filepath, 0, -1, tuple(header))
        return content

    @classmethod
    def write(cls, filepath, header, rows):
        with open(filepath, 'w', newline='') as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(header)
            writer.writerows(rows)

    @classmethod
    def write_from_dict(cls, filepath, header, rows):
        with open(filepath, 'w', newline='') as csv_file:
            writer = csv.DictWriter(csv_file, header)
            writer.writeheader()
            writer.writerows(rows)

    @classmethod
    def append_rows(cls, filepath, rows):
        need_new_line = False
        with open(filepath, 'r') as csv_file:
            content = csv_file.read()
            if not content.endswith('\n'):
                need_new_line = True
        with open(filepath, 'a', newline='') as csv_file:
            writer = csv.writer(csv_file)
            if need_new_line:
                writer.writerow('\n')
            writer.writerows(rows)

    @classmethod
    def insert_rows(cls, filepath, location, rows):
        # location can be negative value, which means from bottom to top
        pass


class ExcelUtil(object):
    @classmethod
    def get_sheet_name(cls,file_path):
        if not os.path.exists(file_path):
            file_path = os.path.join(tc.data_dir, file_path)
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File {file_path} does not exist.")
        with pd.ExcelFile(file_path) as df:
            return df.sheet_names

    @classmethod
    def read_sheet(cls, file_path, sheet_name):
        if not os.path.exists(file_path):
            file_path = os.path.join(tc.data_dir, file_path)
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File {file_path} does not exist.")
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        data_array = df.values.tolist()[1:]
        return data_array

    @classmethod
    def read_filtered_columns(cls, file_path, sheet_name, *columns, header_row=0 , **filters):
        """
        按指定列的组合条件从 Excel 中查找符合条件的行，返回指定列名的值
        :param sheet_name:
        :param file_path: 文件路径
        :param columns: 要读取并返回的数据列，比如：'Error Message', 'Receiver 收货方'
        :param filters: 过滤条件，选出符合条件的行，比如，**{'Authorization 授权': 'ER/PH', 'CFN 产品编号': 'DCB-3535'}
        :return:
        """
        if not os.path.exists(file_path):
            file_path = os.path.join(tc.data_dir, file_path)
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File {file_path} does not exist.")
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
        if len(filters) > 0:
            for key, value in filters.items():
                df = df[df[key] == value]
        if columns is not None and len(columns) > 0:
            df = df[list(columns)]
        return df.to_dict(orient='records')

    @classmethod
    def read_cell(cls, file_path, sheet_name, row_num, column_name, header_row=0):
        """
        按行号和列名读取 Excel 的值
        row_num: 行号，从 1 开始
        """
        if not os.path.exists(file_path):
            file_path = os.path.join(tc.data_dir, file_path)
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File {file_path} does not exist.")
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
        try:
            value = df.at[row_num - 1, column_name]
            return value
        except KeyError:
            raise ValueError(f"Column '{column_name}' does not exist in the Excel file.")
        except IndexError:
            raise ValueError(f"Row '{row_num}' does not exist in the Excel file.")

    @classmethod
    def read_row(cls, file_path, sheet_name, row_num):
        """
        根据行号读取 Excel 文件中的一整行记录

        :param file_path: Excel 文件路径
        :param row_num: 行号，从 1 开始
        :param sheet_name: Excel 工作表名称
        :return: 指定行的记录，作为字典返回
        """
        if not os.path.exists(file_path):
            file_path = os.path.join(tc.data_dir, file_path)
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File {file_path} does not exist.")
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # 检查行号是否在范围内
        if row_num < 1 or row_num > len(df):
            raise IndexError("Row number is out of range.")

        # 获取指定行的记录，并转换为字典
        row_data = df.iloc[row_num - 1].to_dict()

        return row_data


    @classmethod
    def write_to_excel(cls, file_path, sheet_name, data):
        """
        将数据写入 Excel 文件，此方法会覆盖现有的Excel文件
        示例使用，按行组织的数据，使用列表格式
        data = [
            {'Authorization 授权': 'ER/PH', 'CFN 产品编号': 'DCB-3535', 'Error Message': '该批次产品未找到'},
            {'Authorization 授权': 'AAS', 'CFN 产品编号': 'DCB-3536', 'Error Message': '没有授权'},
            {'Authorization 授权': 'AIS', 'CFN 产品编号': 'DCB-3537', 'Error Message': '没有授权'}
        ]
        :param file_path: Excel 文件路径
        :param data: 按行组织的要写入的数据，通常是一个列表，每个元素是一个字典
        :param sheet_name: Excel 工作表名称
        """
        # 如果数据是按行组织的列表，先转换为DataFrame
        if isinstance(data, list):
            data = pd.DataFrame(data)

        # 将 DataFrame 写入 Excel 文件
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            data.to_excel(writer, sheet_name=sheet_name, index=False)

    @classmethod
    def update_excel(cls, file_path, sheet_name, filters, update_data):
        """
        根据条件更新 Excel 文件中的指定记录

        :param file_path: Excel 文件路径
        :param filters: 过滤条件字典，用于查找要更新的行，如：{'Authorization 授权': 'ER/PH', 'CFN 产品编号': 'DCB-3535'}
        :param update_data: 更新的数据字典，如：{'Error Message': '更新后的错误信息'}
        :param sheet_name: Excel 工作表名称
        """
        # 读取现有的 Excel 文件
        if not os.path.exists(file_path):
            file_path = os.path.join(tc.data_dir, file_path)
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File {file_path} does not exist.")
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # 应用过滤条件查找要更新的行
        mask = pd.Series([True] * len(df))  # 初始掩码，所有值为 True
        for key, value in filters.items():
            mask &= (df[key] == value)  # 逐步应用过滤条件

        if mask.sum() == 0:  # 如果没有符合条件的记录
            print("No matching records found to update.")
            return

        # 更新指定的记录
        for key, value in update_data.items():
            df.loc[mask, key] = value  # 仅更新符合条件的行

        # 将更新后的 DataFrame 写回 Excel 文件
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # @classmethod
    # def update_row(cls, file_path, sheet_name, row_num, new_data, header_row=1):
    #     """
    #     根据行号更新 Excel 文件中的一整行记录
    #
    #     :param file_path: Excel 文件路径
    #     :param sheet_name: Excel 工作表名称
    #     :param row_num: 行号，从 1 开始
    #     :param new_data: 新的数据，作为字典传入
    #     """
    #     if not os.path.exists(file_path):
    #         file_path = os.path.join(tc.data_dir, file_path)
    #     if not os.path.exists(file_path):
    #         raise FileNotFoundError(f"File {file_path} does not exist.")
    #
    #     # 读取现有的 Excel 文件
    #     df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row - 1)
    #     # 检查行号是否在范围内
    #     if row_num < 1 or row_num > len(df):
    #         raise IndexError("Row number is out of range.")
    #
    #     # 更新指定行的数据
    #     for key, value in new_data.items():
    #         if key in df.columns:
    #
    #             df.at[row_num-1, key] = value
    #         else:
    #             raise KeyError(f"Column '{key}' does not exist in the Excel file.")
    #
    #     # 将更新后的 DataFrame 写回 Excel 文件，并保留原始格式
    #     book = load_workbook(file_path)
    #     with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    #         writer.sheets.update({ws.title: ws for ws in book.worksheets})
    #         df.to_excel(writer, sheet_name=sheet_name, index=False)

    @classmethod
    def update_row(cls, file_path, sheet_name, row_num, new_data, header_row=1):
        """
        根据行号更新 Excel 文件中的一整行记录

        :param file_path: Excel 文件路径
        :param sheet_name: Excel 工作表名称
        :param row_num: 行号，从 1 开始
        :param new_data: 新的数据，作为字典传入
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File {file_path} does not exist.")

        # 加载现有的工作簿和工作表
        book = load_workbook(file_path)
        sheet = book[sheet_name]

        # 获取标题行
        headers = [cell.value for cell in sheet[header_row]]

        # 检查行号是否在范围内
        if row_num < 1 or row_num > sheet.max_row:
            raise IndexError("Row number is out of range.")

        # 更新指定行的数据
        for key, value in new_data.items():
            if key in headers:
                col_idx = headers.index(key) + 1
                sheet.cell(row=row_num + header_row, column=col_idx, value=value)
            else:
                raise KeyError(f"Column '{key}' does not exist in the Excel file.")

        # 保存更改
        book.save(file_path)

        # Excel上传问题workaround，使用MS Excel打开并保存，缺陷是测试机上必须要装有office
        # todo: 研究openpyxl或者pandas库，更新save方法
        cls.open_and_save_excel(file_path)

    @classmethod
    def append_row(cls, file_path, sheet_name, new_data, header_row=1):
        """
        在当前 Excel 文件的数据行下方追加写入数据，根据列名匹配数据。

        :param file_path: Excel 文件路径
        :param sheet_name: Excel 工作表名称
        :param new_data: 新的数据，作为字典传入
        :param header_row: 列名所在的行
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File {file_path} does not exist.")

        # 加载现有的 Excel 文件
        workbook = load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet {sheet_name} does not exist in the workbook.")
        sheet = workbook[sheet_name]
        # 获取列名和对应的列索引
        headers = {cell.value: cell.column for cell in sheet[header_row]}
        # 准备要追加的新行数据
        new_row = [new_data.get(header, None) for header in headers]
        # 追加新数据到工作表中
        sheet.append(new_row)
        # 保存更改回 Excel 文件
        workbook.save(file_path)

    @classmethod
    def open_and_save_excel(cls, file_path):
        # 获取绝对路径（确保路径正确）
        abs_file_path = os.path.abspath(file_path)

        # 创建 Excel 应用程序对象
        excel = win32.Dispatch('Excel.Application')

        # 可选：设置为不可见，防止弹出 Excel 窗口
        excel.Visible = False
        excel.ScreenUpdating = False

        try:
            # 打开 Excel 文件
            workbook = excel.Workbooks.Open(abs_file_path)

            # 执行保存操作
            workbook.Save()

            # 关闭工作簿
            workbook.Close(SaveChanges=False)
        except Exception as e:
            print(f"An error occurred: {e}")
        finally:
            # 退出 Excel 应用程序
            excel.Quit()

    @classmethod
    def get_hyperlink_list(cls, file_path, sheet_name) -> List:
        """
        从指定的 Excel 工作表中获取所有超链接。

        :param file_path: Excel 文件的路径。
        :param sheet_name: 要访问的工作表名称。
        :return: 包含所有超链接的列表。
        :raises FileNotFoundError: 如果文件路径不存在。
        :raises KeyError: 如果指定的工作表名称不存在。
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File {file_path} does not exist.")
        workbook = load_workbook(file_path)
        value_list=[]
        if sheet_name not in workbook.sheetnames:
            raise KeyError(f"Sheet {sheet_name} does not exist in the workbook.")
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.hyperlink:
                    value_list.append(cell.hyperlink.target)
        return value_list

    @classmethod
    def is_excel_file_empty(cls, file_path):
        workbook = load_workbook(file_path, data_only=True)
        for sheet in workbook.worksheets:
            if any(sheet.iter_rows(values_only=True)):
                return False
        return True


class YamlUtil(object):
    @classmethod
    def read(cls, file_path):
        with open(file_path, mode='r', encoding='utf-8') as fs:
            data = yaml.safe_load(fs)
            return data


class FileUtil(object):
    @classmethod
    def clear_folder(cls,folder_path):
        # 确保文件夹存在
        if not os.path.exists(folder_path):
            print(f"The folder {folder_path} does not exist.")
            return

        # 遍历文件夹中的所有文件和子文件夹
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)

            try:
                # 如果是文件，删除文件
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
                # 如果是文件夹，删除文件夹及其所有内容
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                print(f"Failed to delete {file_path}. Reason: {e}")

    @classmethod
    def get_recent_download_file_name(cls,file_path,time_to_wait_for_download_complete = 1): #动态等待用不了时换下面的等待方法
        start = datetime.now()
        former_files = [file for file in os.listdir(file_path) if not file.endswith(('.tmp', '.part', '.crdownload'))]
        while datetime.now() - start < timedelta(seconds=time_to_wait_for_download_complete):
            time.sleep(0.1)
            files = [file for file in os.listdir(file_path) if not file.endswith(('.tmp', '.part', '.crdownload'))]
            if len(files) > len(former_files):
                files.sort(key=lambda x: os.path.getmtime(os.path.join(file_path, x)), reverse=True)
                latest_file = files[0]
                return latest_file
        raise FileNotFoundError("File does not exist.")

        # time.sleep(time_to_wait_for_download_complete)
        # files = os.listdir(file_path)
        # files.sort(key=lambda x: os.path.getmtime(os.path.join(file_path, x)))
        # latest_file = files[0]
        # return latest_file

